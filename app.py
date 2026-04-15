# coding=utf-8
import os, io, uuid, threading, json, hashlib, copy
from flask import Flask, request, jsonify, send_file, render_template
from flask_socketio import SocketIO, join_room, leave_room, emit
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_GENERAL
from converter import krutidev_to_unicode
import db as DB

# ── App setup ─────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.config['SECRET_KEY']        = os.environ.get('SECRET_KEY', 'kd-editor-secret-2024')
app.config['UPLOAD_FOLDER']     = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

try:
    import eventlet  # noqa: F401
    _async_mode = 'eventlet'
except ImportError:
    _async_mode = 'threading'

socketio = SocketIO(app, cors_allowed_origins='*', async_mode=_async_mode)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
progress_store = {}

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
DB.init_db()


def allowed_file(f):
    return '.' in f and f.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _needs_conversion(text: str) -> bool:
    dev = sum(1 for c in text if '\u0900' <= c <= '\u097F')
    if dev > len(text) * 0.3: return False
    non_alpha = sum(1 for c in text if not c.isalpha())
    return non_alpha != len(text)


# ── Excel helpers ─────────────────────────────────────────────────────────────

# In-memory store mapping file_id → original upload path (for format preservation)
_source_files: dict = {}


def _copy_cell_style(src_cell, dst_cell):
    """Copy all style attributes from src to dst cell."""
    if src_cell.has_style:
        dst_cell.font      = copy.copy(src_cell.font)
        dst_cell.fill      = copy.copy(src_cell.fill)
        dst_cell.border    = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection    = copy.copy(src_cell.protection)


def excel_to_sheets(path: str) -> tuple[dict, list]:
    wb = load_workbook(path, data_only=True)
    sheets_data, sheet_names = {}, wb.sheetnames

    # Legacy Krutidev font names — cells using these fonts store Krutidev-encoded text
    KRUTIDEV_FONTS = {
        'kruti dev', 'krutidev', 'kruti', 'devlys', 'chanakya', 'shusha',
        'walkman chanakya', 'dvb-ttsurekh', 'dvb-ttyogesh',
    }

    def _is_krutidev_font(font_name: str) -> bool:
        if not font_name:
            return False
        fn = font_name.lower().strip()
        return any(kf in fn for kf in KRUTIDEV_FONTS)

    for name in sheet_names:
        ws = wb[name]
        rows, col_widths, row_heights = [], {}, {}
        cell_fonts = {}

        # ── Read merged cell ranges from the workbook ──────────────────────
        # openpyxl stores merges as "A1:C3" style strings
        merge_list = []
        for merge_range in ws.merged_cells.ranges:
            merge_list.append({
                'row':     merge_range.min_row - 1,   # 0-based
                'col':     merge_range.min_col - 1,
                'rowspan': merge_range.max_row - merge_range.min_row + 1,
                'colspan': merge_range.max_col - merge_range.min_col + 1,
            })

        for row in ws.iter_rows():
            rd = []
            row_idx = row[0].row if row else 1
            for cell in row:
                val = cell.value
                val = '' if val is None else (str(val) if not isinstance(val, str) else val)
                rd.append(val)

                cl = get_column_letter(cell.column)
                if cl not in col_widths:
                    cd = ws.column_dimensions.get(cl)
                    if cd and cd.width and cd.width > 0:
                        col_widths[cl] = max(40, round(cd.width * 8))
                    else:
                        col_widths[cl] = 100

                if cell.font and cell.font.name:
                    font_name = cell.font.name
                    col_idx = cell.column - 1
                    cell_fonts[f'{row_idx - 1},{col_idx}'] = font_name

            rd_obj = ws.row_dimensions.get(row_idx)
            if rd_obj and rd_obj.height and rd_obj.height > 0:
                row_heights[str(row_idx)] = round(rd_obj.height * 1.33)

            rows.append(rd)

        mc = max((len(r) for r in rows), default=0)
        rows = [r + [''] * (mc - len(r)) for r in rows]

        krutidev_cells = [k for k, fn in cell_fonts.items() if _is_krutidev_font(fn)]

        sheets_data[name] = {
            'rows': rows,
            'col_widths': col_widths,
            'row_heights': row_heights,
            'cell_fonts': cell_fonts,
            'krutidev_cells': krutidev_cells,
            'merge_list': merge_list,   # ← actual merges from the file
            'row_count': len(rows),
            'col_count': mc,
        }
    wb.close()
    return sheets_data, sheet_names


def sheets_to_excel(sheets: dict, convert_mode=False, source_path: str = None) -> io.BytesIO:
    """
    Build an Excel workbook from sheets data.
    If source_path is provided, copy all cell formatting from the original file
    so the output preserves column widths, row heights, fonts, fills, borders, etc.
    Only cell *values* are replaced with the edited/converted content.
    """
    if source_path and os.path.exists(source_path):
        # Load original as formatting template
        wb = load_workbook(source_path)
        for name, sd in sheets.items():
            if name in wb.sheetnames:
                ws = wb[name]
            else:
                ws = wb.create_sheet(title=name[:31])

            new_rows = sd.get('rows', [])
            for ri, row_data in enumerate(new_rows, 1):
                for ci, val in enumerate(row_data, 1):
                    cell = ws.cell(row=ri, column=ci)
                    if convert_mode and isinstance(val, str) and val.strip():
                        cell.value = krutidev_to_unicode(val)
                    else:
                        cell.value = val if val != '' else None

            # Apply merge cells from editor state
            for mc in sd.get('merge_cells', []):
                try:
                    ws.merge_cells(
                        start_row=mc['row']+1, start_column=mc['col']+1,
                        end_row=mc['row']+mc['rowspan'],
                        end_column=mc['col']+mc['colspan']
                    )
                except Exception:
                    pass

            # Remove extra rows that were in the original but not in new data
            max_new_row = len(new_rows)
            if ws.max_row > max_new_row:
                for ri in range(max_new_row + 1, ws.max_row + 1):
                    for ci in range(1, ws.max_column + 1):
                        ws.cell(row=ri, column=ci).value = None

        # Remove sheets that no longer exist
        for sname in list(wb.sheetnames):
            if sname not in sheets:
                del wb[sname]
    else:
        # Fallback: create fresh workbook (no formatting)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for name, sd in sheets.items():
            ws = wb.create_sheet(title=name[:31])
            for ri, row in enumerate(sd.get('rows', []), 1):
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(row=ri, column=ci)
                    if convert_mode and isinstance(val, str) and val.strip():
                        cell.value = krutidev_to_unicode(val)
                    else:
                        cell.value = val if val != '' else None
            # Apply merge cells
            for mc in sd.get('merge_cells', []):
                try:
                    ws.merge_cells(
                        start_row=mc['row']+1, start_column=mc['col']+1,
                        end_row=mc['row']+mc['rowspan'],
                        end_column=mc['col']+mc['colspan']
                    )
                except Exception:
                    pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def convert_excel_job(input_path, output_path, job_id, mode='ktu'):
    """
    mode: 'ktu' = Krutidev → Unicode (default)
          'utk' = Unicode → Krutidev
    """
    try:
        from converter import unicode_to_krutidev
        progress_store[job_id] = {'status': 'processing', 'progress': 0, 'mode': mode}
        wb = load_workbook(input_path)
        sheets = wb.sheetnames
        for si, sn in enumerate(sheets):
            ws = wb[sn]
            rows = list(ws.iter_rows())
            for ri, row in enumerate(rows):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        v = cell.value.strip()
                        if not v:
                            continue
                        if mode == 'ktu':
                            if _needs_conversion(v):
                                cell.value = krutidev_to_unicode(cell.value)
                        else:  # utk — convert any Devanagari Unicode text back to Krutidev
                            has_devanagari = any('\u0900' <= c <= '\u097F' for c in v)
                            if has_devanagari:
                                cell.value = unicode_to_krutidev(cell.value)
                pct = int(((si / len(sheets)) + (ri / len(rows) / len(sheets))) * 100)
                progress_store[job_id]['progress'] = min(pct, 95)
        wb.save(output_path)
        progress_store[job_id] = {'status': 'done', 'progress': 100, 'output': output_path, 'mode': mode}
    except Exception as e:
        progress_store[job_id] = {'status': 'error', 'message': str(e)}


# ── Pages ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/editor')
def editor_new():
    return render_template('editor.html', doc_id=None, access='edit', title='Untitled')


@app.route('/sheet/<doc_id>')
def sheet_view(doc_id):
    doc = DB.get_doc(doc_id)
    if not doc:
        return render_template('404.html'), 404
    return render_template('editor.html',
                           doc_id=doc_id,
                           access=doc['access'],
                           title=doc['title'])


@app.route('/font/download')
def font_download():
    """Serve KrutiDev010.ttf as a download."""
    font_path = os.path.join(app.root_path, 'static', 'fonts', 'KrutiDev010.ttf')
    if not os.path.exists(font_path):
        return jsonify({'error': 'Font file not found'}), 404
    return send_file(font_path, as_attachment=True,
                     download_name='KrutiDev010.ttf',
                     mimetype='font/truetype')


@app.route('/sitemap.xml')
def sitemap():
    """XML sitemap for SEO crawlers."""
    from flask import Response
    xml = '''<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9"
        xmlns:xhtml="http://www.w3.org/1999/xhtml">
  <url>
    <loc>https://krutidevstudio.com/</loc>
    <changefreq>weekly</changefreq>
    <priority>1.0</priority>
    <xhtml:link rel="alternate" hreflang="hi" href="https://krutidevstudio.com/"/>
    <xhtml:link rel="alternate" hreflang="en" href="https://krutidevstudio.com/"/>
  </url>
  <url>
    <loc>https://krutidevstudio.com/editor</loc>
    <changefreq>monthly</changefreq>
    <priority>0.8</priority>
  </url>
</urlset>'''
    return Response(xml, mimetype='application/xml')


@app.route('/robots.txt')
def robots():
    """robots.txt for SEO crawlers."""
    from flask import Response
    txt = '''User-agent: *
Allow: /
Disallow: /api/
Disallow: /uploads/
Disallow: /progress/
Disallow: /download/

Sitemap: https://krutidevstudio.com/sitemap.xml
'''
    return Response(txt, mimetype='text/plain')


# ── Legacy converter routes ───────────────────────────────────────────────────

@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if not file.filename or not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type'}), 400
    mode = request.form.get('mode', 'ktu')   # 'ktu' or 'utk'
    if mode not in ('ktu', 'utk'):
        mode = 'ktu'
    filename = secure_filename(file.filename)
    job_id = str(uuid.uuid4())
    inp = os.path.join(app.config['UPLOAD_FOLDER'], f'{job_id}_input_{filename}')
    out = os.path.join(app.config['UPLOAD_FOLDER'], f'{job_id}_output_{filename}')
    file.save(inp)
    try:
        load_workbook(inp, read_only=True).close()
    except Exception:
        os.remove(inp)
        return jsonify({'error': 'Invalid or corrupted Excel file'}), 400
    t = threading.Thread(target=convert_excel_job, args=(inp, out, job_id, mode), daemon=True)
    t.start()
    return jsonify({'job_id': job_id, 'filename': filename, 'mode': mode})


@app.route('/progress/<job_id>')
def progress(job_id):
    d = progress_store.get(job_id)
    return jsonify(d) if d else (jsonify({'status': 'not_found'}), 404)


@app.route('/download/<job_id>')
def download(job_id):
    d = progress_store.get(job_id)
    if not d or d.get('status') != 'done':
        return jsonify({'error': 'File not ready'}), 404
    op = d.get('output')
    if not op or not os.path.exists(op):
        return jsonify({'error': 'Output file not found'}), 404
    bn = os.path.basename(op)
    orig = bn.replace(f'{job_id}_output_', '')
    name, ext = os.path.splitext(orig)
    mode = d.get('mode', 'ktu')
    suffix = '_unicode' if mode == 'ktu' else '_krutidev'

    # Read file into memory so we can delete it before sending
    with open(op, 'rb') as fh:
        file_bytes = fh.read()

    # Clean up both input and output files + remove from progress store
    inp = op.replace(f'{job_id}_output_', f'{job_id}_input_')
    for path in (op, inp):
        try:
            os.remove(path)
        except OSError:
            pass
    progress_store.pop(job_id, None)

    return send_file(
        io.BytesIO(file_bytes),
        as_attachment=True,
        download_name=f'{name}{suffix}{ext}',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ── Editor API ────────────────────────────────────────────────────────────────

@app.route('/api/convert-text', methods=['POST'])
def api_convert_text():
    """Convert text between Krutidev and Unicode."""
    try:
        p = request.get_json(force=True)
        text = p.get('text', '')
        mode = p.get('mode', 'ktu')  # ktu = krutidev→unicode, utk = unicode→krutidev
        if mode == 'ktu':
            result = krutidev_to_unicode(text)
        else:
            # Unicode→Krutidev: import reverse converter
            from converter import unicode_to_krutidev
            result = unicode_to_krutidev(text)
        return jsonify({'result': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/load', methods=['POST'])
def api_load():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if not file.filename or not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type'}), 400
    filename = secure_filename(file.filename)
    fid = str(uuid.uuid4())
    path = os.path.join(app.config['UPLOAD_FOLDER'], f'{fid}_{filename}')
    file.save(path)
    try:
        sheets_data, sheet_names = excel_to_sheets(path)
        # Remember the source file for format-preserving export
        _source_files[fid] = path
        # Evict oldest entries if store grows too large (keep last 50)
        if len(_source_files) > 50:
            oldest = list(_source_files.keys())[0]
            _source_files.pop(oldest, None)
        return jsonify({'file_id': fid, 'filename': filename,
                        'sheets': sheets_data, 'sheet_names': sheet_names})
    except Exception as e:
        try:
            os.remove(path)
        except OSError:
            pass
        return jsonify({'error': str(e)}), 500


@app.route('/api/save', methods=['POST'])
def api_save():
    try:
        p = request.get_json(force=True)
        file_id = p.get('file_id')
        source_path = _source_files.get(file_id) if file_id else None
        buf = sheets_to_excel(p.get('sheets', {}), p.get('convert', False), source_path)
        name = p.get('filename', 'edited').rsplit('.', 1)[0]
        suffix = '_unicode' if p.get('convert') else '_edited'
        return send_file(buf, as_attachment=True, download_name=f'{name}{suffix}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Share / Document API ──────────────────────────────────────────────────────

@app.route('/api/share', methods=['POST'])
def api_share():
    """Create or update a shared document. Returns shareable URL."""
    try:
        p = request.get_json(force=True)
        sheets  = p.get('sheets', {})
        names   = p.get('sheet_names', list(sheets.keys()))
        title   = p.get('title', 'Untitled')
        access  = p.get('access', 'edit')   # 'view' | 'edit'
        pwd     = p.get('password')
        doc_id  = p.get('doc_id')           # if updating existing
        file_id = p.get('file_id')          # source file for format preservation

        pwd_hash = hashlib.sha256(pwd.encode()).hexdigest() if pwd else None

        # Resolve source path and store it in the doc metadata
        source_path = _source_files.get(file_id) if file_id else None

        if doc_id:
            DB.update_doc(doc_id, sheets, names, title, source_path=source_path)
        else:
            doc_id = DB.create_doc(title, sheets, names, access, pwd_hash,
                                   source_path=source_path)

        url = request.host_url.rstrip('/') + f'/sheet/{doc_id}'
        return jsonify({'doc_id': doc_id, 'url': url, 'access': access})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/doc/<doc_id>', methods=['GET'])
def api_get_doc(doc_id):
    """Load a shared document's data."""
    doc = DB.get_doc(doc_id)
    if not doc:
        return jsonify({'error': 'Document not found'}), 404
    pwd = request.args.get('pwd')
    if doc['password']:
        if not pwd or hashlib.sha256(pwd.encode()).hexdigest() != doc['password']:
            return jsonify({'error': 'Password required', 'protected': True}), 403
    sheets = doc['data']
    names  = doc['sheet_names']
    # Ensure we always return at least one sheet
    if not sheets or not names:
        sheets = {}
        names  = []
    return jsonify({'doc_id': doc_id, 'title': doc['title'],
                    'sheets': sheets, 'sheet_names': names,
                    'access': doc['access'], 'updated_at': doc['updated_at']})


@app.route('/api/doc/<doc_id>', methods=['DELETE'])
def api_delete_doc(doc_id):
    DB.delete_doc(doc_id)
    return jsonify({'ok': True})


@app.route('/api/docs', methods=['GET'])
def api_list_docs():
    return jsonify({'docs': DB.list_docs()})


@app.route('/api/doc/<doc_id>/download', methods=['GET'])
def api_download_doc(doc_id):
    doc = DB.get_doc(doc_id)
    if not doc:
        return jsonify({'error': 'Not found'}), 404
    convert = request.args.get('convert', 'false').lower() == 'true'
    source_path = doc.get('source_path')
    buf = sheets_to_excel(doc['data'], convert, source_path)
    name = doc['title'].replace(' ', '_')
    suffix = '_unicode' if convert else ''
    return send_file(buf, as_attachment=True, download_name=f'{name}{suffix}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── WebSocket collaboration ───────────────────────────────────────────────────

# Track active users per room: { doc_id: { sid: {name, color} } }
_rooms: dict[str, dict] = {}
_COLORS = ['#6366f1','#10b981','#f59e0b','#ef4444','#8b5cf6','#06b6d4','#f97316','#ec4899']


def _room_users(doc_id):
    return list(_rooms.get(doc_id, {}).values())


@socketio.on('join')
def on_join(data):
    doc_id = data.get('doc_id')
    name   = data.get('name', 'Anonymous')
    if not doc_id:
        return
    join_room(doc_id)
    _rooms.setdefault(doc_id, {})
    color = _COLORS[len(_rooms[doc_id]) % len(_COLORS)]
    _rooms[doc_id][request.sid] = {'name': name, 'color': color, 'sid': request.sid}
    emit('users_update', {'users': _room_users(doc_id)}, to=doc_id)


@socketio.on('disconnect')
def on_disconnect():
    sid = request.sid
    for doc_id, users in list(_rooms.items()):
        if sid in users:
            del users[sid]
            emit('users_update', {'users': _room_users(doc_id)}, to=doc_id)
            if not users:
                del _rooms[doc_id]
            break


@socketio.on('cell_change')
def on_cell_change(data):
    """Broadcast a cell edit to all other users in the room."""
    doc_id = data.get('doc_id')
    if not doc_id:
        return
    # Persist the change to DB (last-write-wins)
    doc = DB.get_doc(doc_id)
    if doc and doc['access'] == 'edit':
        sheet  = data.get('sheet')
        row    = data.get('row')
        col    = data.get('col')
        value  = data.get('value', '')
        sheets = doc['data']
        if sheet in sheets and row < len(sheets[sheet]['rows']):
            while col >= len(sheets[sheet]['rows'][row]):
                sheets[sheet]['rows'][row].append('')
            sheets[sheet]['rows'][row][col] = value
            DB.update_doc(doc_id, sheets, doc['sheet_names'])
    # Broadcast to everyone else in the room
    emit('cell_change', data, to=doc_id, include_self=False)


@socketio.on('cursor_move')
def on_cursor(data):
    doc_id = data.get('doc_id')
    if doc_id:
        user = _rooms.get(doc_id, {}).get(request.sid, {})
        emit('cursor_move', {**data, 'user': user}, to=doc_id, include_self=False)


@socketio.on('title_change')
def on_title(data):
    doc_id = data.get('doc_id')
    title  = data.get('title', 'Untitled')
    if doc_id:
        doc = DB.get_doc(doc_id)
        if doc:
            DB.update_doc(doc_id, doc['data'], doc['sheet_names'], title)
        emit('title_change', {'title': title}, to=doc_id, include_self=False)


@app.errorhandler(413)
def too_large(e):
    return jsonify({'error': 'File too large. Max 50MB'}), 413


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8765))
    socketio.run(app, debug=False, host='0.0.0.0', port=port)
